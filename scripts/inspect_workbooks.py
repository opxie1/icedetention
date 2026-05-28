"""Print sheet and header layout for every workbook in a directory."""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook


def _peek_sheet(sheet, max_scan: int = 15) -> dict:
    rows = []
    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        rows.append(row)
        if i >= max_scan:
            break

    best_idx, best_score = 0, -1
    for idx, row in enumerate(rows):
        cells = [c for c in row if c is not None and str(c).strip() != ""]
        score = len(cells) if len(cells) >= 5 else 0
        if score > best_score:
            best_idx, best_score = idx, score

    header_row_num = best_idx + 1 if best_score > 0 else None
    header_values = (
        [str(c).strip() if c is not None else "" for c in rows[best_idx]]
        if header_row_num is not None
        else []
    )

    sample_row = []
    if header_row_num is not None and best_idx + 1 < len(rows):
        sample_row = [
            str(c).strip() if c is not None else ""
            for c in rows[best_idx + 1]
        ]

    return {
        "max_col": max(len(r) for r in rows) if rows else 0,
        "header_row": header_row_num,
        "headers": header_values,
        "sample": sample_row,
        "preamble": [
            (i + 1, [str(c).strip() if c is not None else "" for c in r if c is not None])
            for i, r in enumerate(rows[: best_idx if best_idx else 5])
            if any(c is not None and str(c).strip() != "" for c in r)
        ],
    }


def inspect(path: Path) -> None:
    print("=" * 100)
    print(f"FILE: {path.name}")
    print(f"SIZE: {path.stat().st_size / 1_000_000:.1f} MB")
    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception as e:
        print(f"  ERROR opening: {e}")
        return
    try:
        print(f"SHEETS: {wb.sheetnames}")
        for sname in wb.sheetnames:
            sheet = wb[sname]
            info = _peek_sheet(sheet)
            print(f"\n  --- sheet '{sname}' ---")
            print(f"    max_columns_in_first_15_rows: {info['max_col']}")
            if info["preamble"]:
                print(f"    preamble (rows above header):")
                for row_num, vals in info["preamble"]:
                    short = vals[:3]
                    suffix = " ..." if len(vals) > 3 else ""
                    print(f"      row {row_num}: {short}{suffix}")
            print(f"    detected_header_row: {info['header_row']}")
            if info["headers"]:
                print(f"    n_headers: {len(info['headers'])}")
                for i, h in enumerate(info["headers"], start=1):
                    print(f"      col {i:>2}: {h!r}")
            if info["sample"]:
                print(f"    sample data row (truncated to 200 chars):")
                line = " | ".join(info["sample"])
                print(f"      {line[:200]}")
    finally:
        wb.close()


def main(argv: list[str]) -> int:
    if len(argv) != 2:
        print(__doc__)
        return 2
    folder = Path(argv[1])
    if not folder.is_dir():
        print(f"Not a directory: {folder}", file=sys.stderr)
        return 2
    files = sorted(folder.glob("*.xlsx"))
    print(f"Found {len(files)} workbook(s) in {folder}")
    for f in files:
        inspect(f)
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
