"""Stream FOIA xlsx workbooks into slim CSVs."""

from __future__ import annotations

import csv
import datetime as dt
import gzip
import logging
import re
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook

from . import config

log = logging.getLogger(__name__)


def _coerce_date(value) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        if 1 <= float(value) <= 80000:
            try:
                base = dt.datetime(1899, 12, 30)
                return (base + dt.timedelta(days=float(value))).date().isoformat()
            except (OverflowError, ValueError):
                return ""
        return ""
    return str(value).strip()


def _coerce_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def fiscal_year_from_filename(path: Path) -> int | None:
    m = re.search(config.INPUT_FILENAME_RE, path.name)
    if not m:
        return None
    raw = m.group(1)
    fy = int(raw)
    if len(raw) == 4:
        return fy
    return 2000 + fy if fy < 50 else 1900 + fy


def fiscal_year_from_sheet(sheet_name: str) -> int | None:
    m = re.match(r"^FY(\d{4})$", sheet_name or "")
    return int(m.group(1)) if m else None


def _validate_header(row_values: list) -> None:
    actual = [_coerce_text(v) for v in row_values[: len(config.EXPECTED_HEADERS)]]
    if actual != config.EXPECTED_HEADERS:
        mismatches = [
            (i + 1, exp, got)
            for i, (exp, got) in enumerate(zip(config.EXPECTED_HEADERS, actual))
            if exp != got
        ]
        msg = "Header row does not match expected layout.\n"
        for col, exp, got in mismatches[:10]:
            msg += f"  col {col}: expected {exp!r}, got {got!r}\n"
        raise ValueError(msg)


def _iter_data_rows(sheet) -> Iterator[tuple]:
    cols = (
        config.COL_DETENTION_BOOK_IN,
        config.COL_DETENTION_FACILITY,
        config.COL_DETENTION_FAC_CODE,
        config.COL_DETENTION_BOOK_OUT,
        config.COL_ANON_ID,
    )
    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if i < config.HEADER_ROW:
            continue
        if i == config.HEADER_ROW:
            _validate_header(list(row))
            continue
        if all(v is None for v in row):
            continue
        yield tuple(row[c - 1] if c - 1 < len(row) else None for c in cols)


def extract_workbook(xlsx_path: Path, out_path: Path, gzip_output: bool = True) -> dict:
    log.info("opening %s", xlsx_path.name)
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    try:
        target = next(
            (s for s in wb.sheetnames if re.match(config.SHEET_NAME_RE, s)),
            None,
        )
        if target is None:
            raise ValueError(
                f"{xlsx_path.name}: no sheet named FY<YYYY> "
                f"(found: {wb.sheetnames!r})"
            )
        sheet = wb[target]
        fy = fiscal_year_from_sheet(target) or fiscal_year_from_filename(xlsx_path)
        if fy is None:
            raise ValueError(f"{xlsx_path.name}: cannot resolve fiscal year")

        out_path.parent.mkdir(parents=True, exist_ok=True)
        opener = gzip.open if gzip_output else open
        mode = "wt"
        rows_written = 0
        rows_skipped = 0
        with opener(out_path, mode, encoding="utf-8", newline="") as fh:
            writer = csv.writer(fh)
            writer.writerow(config.EXTRACT_COLUMNS)
            for book_in, facility, code, book_out, person_id in _iter_data_rows(sheet):
                facility_s = _coerce_text(facility)
                code_s = _coerce_text(code)
                if not facility_s and not code_s:
                    rows_skipped += 1
                    continue
                writer.writerow(
                    [
                        fy,
                        facility_s,
                        code_s,
                        _coerce_date(book_in),
                        _coerce_date(book_out),
                        _coerce_text(person_id),
                    ]
                )
                rows_written += 1
                if rows_written % 50000 == 0:
                    log.info("  ... %s rows", f"{rows_written:,}")
        log.info(
            "wrote %s rows (skipped %s) -> %s",
            f"{rows_written:,}",
            f"{rows_skipped:,}",
            out_path.name,
        )
        return {
            "fiscal_year": fy,
            "rows_written": rows_written,
            "rows_skipped": rows_skipped,
            "sheet": target,
            "input": str(xlsx_path),
            "output": str(out_path),
        }
    finally:
        wb.close()


def discover_inputs(input_dir: Path) -> list[Path]:
    candidates = sorted(input_dir.glob("*Detentions_FY*.xlsx"))
    return [
        p for p in candidates
        if not p.name.startswith("~$")
        and fiscal_year_from_filename(p) is not None
    ]


def _encounters_period_tag(sheet_name: str) -> str:
    s = (sheet_name or "").lower()
    if "<" in s and "10012024" in s:
        return "pre_20241001"
    if ">=" in s and "10012024" in s:
        return "from_20241001"
    safe = re.sub(r"[^a-z0-9]+", "_", s).strip("_")
    return safe or "unknown"


def _validate_encounters_header(row_values: list) -> None:
    actual = [_coerce_text(v) for v in row_values[: len(config.ENCOUNTERS_EXPECTED_HEADERS)]]
    if actual != config.ENCOUNTERS_EXPECTED_HEADERS:
        mismatches = [
            (i + 1, exp, got)
            for i, (exp, got) in enumerate(zip(config.ENCOUNTERS_EXPECTED_HEADERS, actual))
            if exp != got
        ]
        msg = "Encounters header row does not match expected layout.\n"
        for col, exp, got in mismatches[:10]:
            msg += f"  col {col}: expected {exp!r}, got {got!r}\n"
        raise ValueError(msg)


def _iter_encounter_rows(sheet) -> Iterator[tuple]:
    cols = (
        config.ENC_COL_EVENT_DATE,
        config.ENC_COL_RESPONSIBLE_AOR,
        config.ENC_COL_RESPONSIBLE_SITE,
        config.ENC_COL_LEAD_EVENT_TYPE,
        config.ENC_COL_EVENT_TYPE,
        config.ENC_COL_FINAL_PROGRAM,
        config.ENC_COL_FINAL_PROGRAM_GROUP,
        config.ENC_COL_PROCESSING_DISP,
        config.ENC_COL_UNIQUE_ID,
    )
    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if i < config.ENCOUNTERS_HEADER_ROW:
            continue
        if i == config.ENCOUNTERS_HEADER_ROW:
            _validate_encounters_header(list(row))
            continue
        if all(v is None for v in row):
            continue
        yield tuple(row[c - 1] if c - 1 < len(row) else None for c in cols)


def discover_encounters(input_dir: Path) -> list[Path]:
    candidates = sorted(input_dir.glob(config.ENCOUNTERS_FILENAME_GLOB))
    return [p for p in candidates if not p.name.startswith("~$")]


def extract_encounters_workbook(
    xlsx_path: Path, out_path: Path, gzip_output: bool = True
) -> dict:
    log.info("opening %s (encounters)", xlsx_path.name)
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    try:
        target_sheets = [
            s for s in wb.sheetnames
            if s.startswith(config.ENCOUNTERS_SHEET_PREFIX)
        ]
        if not target_sheets:
            raise ValueError(
                f"{xlsx_path.name}: no sheets named {config.ENCOUNTERS_SHEET_PREFIX!r}* "
                f"(found: {wb.sheetnames!r})"
            )

        out_path.parent.mkdir(parents=True, exist_ok=True)
        opener = gzip.open if gzip_output else open
        rows_written = 0
        rows_skipped = 0
        per_sheet: dict[str, int] = {}
        with opener(out_path, "wt", encoding="utf-8", newline="") as fh:
            writer = csv.writer(fh)
            writer.writerow(config.ENCOUNTERS_EXTRACT_COLUMNS)
            for sname in target_sheets:
                tag = _encounters_period_tag(sname)
                log.info("  sheet %r -> tag %r", sname, tag)
                sheet = wb[sname]
                sheet_count = 0
                for (
                    event_date, aor, site, lead_event_type, event_type,
                    final_program, final_program_group, processing_disp,
                    person_id,
                ) in _iter_encounter_rows(sheet):
                    site_s = _coerce_text(site)
                    if not site_s:
                        rows_skipped += 1
                        continue
                    writer.writerow([
                        tag,
                        _coerce_date(event_date),
                        _coerce_text(aor),
                        site_s,
                        _coerce_text(lead_event_type),
                        _coerce_text(event_type),
                        _coerce_text(final_program),
                        _coerce_text(final_program_group),
                        _coerce_text(processing_disp),
                        _coerce_text(person_id),
                    ])
                    rows_written += 1
                    sheet_count += 1
                    if rows_written % 50000 == 0:
                        log.info("    ... %s rows", f"{rows_written:,}")
                per_sheet[sname] = sheet_count
        log.info(
            "wrote %s rows (skipped %s) -> %s",
            f"{rows_written:,}",
            f"{rows_skipped:,}",
            out_path.name,
        )
        return {
            "rows_written": rows_written,
            "rows_skipped": rows_skipped,
            "per_sheet": per_sheet,
            "input": str(xlsx_path),
            "output": str(out_path),
        }
    finally:
        wb.close()
